import pandas as pd
import numpy as np
import keras_tuner
import pickle
import shap
import tensorflow as tf
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from keras import layers
from pathlib import Path
from tensorflow import keras
from sklearn.metrics import mean_squared_error
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

np.random.seed(42)
tf.random.set_seed(42)

gpus = tf.config.experimental.list_physical_devices(device_type='GPU')
for gpu in gpus:
    tf.config.experimental.set_memory_growth(gpu, True)

select_model = 3  
search_time = 15
forecast_period = 5
steps = 24
search_project_name = "{}_{}_{}".format(search_time, forecast_period, steps)
save_search_path = Path('E:/draft/00') / search_project_name
save_search_path.mkdir(exist_ok=True)
save_results_path = Path('E:/draft') / search_project_name
save_results_path.mkdir(exist_ok=True)


data = pd.read_excel('北江数据(13降雨+4流量+输入步长48h).xlsx') 
rainfall_data = data.iloc[:, 1:18].values 
flow_data = data.iloc[:, 20].values 
flow_data_rank_conversion = flow_data.reshape(-1, 1)


if select_model == 1:
    def build_model(hp):
        inputs = keras.Input(shape=(steps, 17))
        z = layers.Conv1D(filters=hp.Choice('conv1_filters', [8, 16, 32, 64]), kernel_size=3, activation='relu')(inputs)
        # z = layers.MaxPooling1D(pool_size=2)(z)
        # z = layers.Conv1D(filters=hp.Choice('conv2_filters', [8, 16, 32, 64]), kernel_size=3,activation='relu')(z)
        # z = layers.MaxPooling1D(pool_size=2)(z)
        # z = layers.LSTM(units=hp.Choice('lstm_units', [8, 16, 32, 64]))(z)
        # z = layers.LeakyReLU()(z)
        z = layers.Flatten()(z)
        z = layers.Dense(units=hp.Choice('dense_units', [32, 64, 128]), kernel_initializer='random_normal')(z)
        z = layers.Dropout(hp.Float('dropout', min_value=0.0, max_value=0.5, step=0.1))(z)
        outputs = layers.Dense(forecast_period)(z)
        model = keras.Model(inputs, outputs)
        model.compile(optimizer=keras.optimizers.Adam(0.001), loss='mse')
        return model


    tuner = keras_tuner.RandomSearch(build_model, objective='val_loss', max_trials=15,
                                     directory=save_search_path, project_name=search_project_name)
    tuner.search(train_ds, epochs=40, validation_data=validate_ds)
    best_model = tuner.get_best_models(num_models=1)[0]
    best_hyperparameters = tuner.get_best_hyperparameters(num_trials=1)[0]
    best_model.compile(optimizer=keras.optimizers.Adam(0.001), loss='mse')
    history = best_model.fit(train_ds, epochs=40, verbose=1, validation_data=validate_ds)
    val_loss_avg_mse = best_model.evaluate(validate_ds, verbose=1)
    print(f'Average MSE: {val_loss_avg_mse}')
    print('Best conv1_filters:', best_hyperparameters.get('conv1_filters'))
    print('Best dense_units:', best_hyperparameters.get('dense_units'))
    print('Best dropout:', best_hyperparameters.get('dropout'))

elif select_model == 2:
    def build_model(hp):
        inputs = keras.Input(shape=(steps, 17))
        z = layers.LSTM(units=hp.Choice('lstm_units', [8, 16, 32, 64, 128]))(inputs)
        z = layers.LeakyReLU()(z)
        z = layers.Flatten()(z)
        z = layers.Dense(units=hp.Choice('dense_units', [32, 64, 128]), kernel_initializer='random_normal')(z)
        z = layers.Dropout(hp.Float('dropout', min_value=0.0, max_value=0.5, step=0.1))(z)
        outputs = layers.Dense(forecast_period)(z)
        model = keras.Model(inputs, outputs)
        model.compile(optimizer=keras.optimizers.Adam(0.001), loss='mse')
        return model


    tuner = keras_tuner.RandomSearch(build_model, objective='val_loss', max_trials=15,
                                     directory=save_search_path, project_name=search_project_name)
    tuner.search(train_ds, epochs=40, validation_data=validate_ds)
    best_model = tuner.get_best_models(num_models=1)[0]
    best_hyperparameters = tuner.get_best_hyperparameters(num_trials=1)[0]
    best_model.compile(optimizer=keras.optimizers.Adam(0.001), loss='mse')
    history = best_model.fit(train_ds, epochs=40, verbose=1, validation_data=validate_ds)
    val_loss_avg_mse = best_model.evaluate(validate_ds, verbose=1)
    print(f'Average MSE: {val_loss_avg_mse}')
    print('Best lstm_units:', best_hyperparameters.get('lstm_units'))
    print('Best dense_units:', best_hyperparameters.get('dense_units'))
    print('Best dropout:', best_hyperparameters.get('dropout'))

elif select_model == 3:
    if forecast_period == 1:
        def build_model(hp):
            inputs = keras.Input(shape=(steps, 17))
            z = layers.Conv1D(filters=hp.Choice('conv1_filters', [8, 16, 32, 64, 128]), kernel_size=3, activation='relu')(inputs)
            z = layers.LSTM(units=hp.Choice('lstm1_units', [8, 16, 32, 64]))(z)
            z = layers.LeakyReLU()(z)
            z = layers.Flatten()(z)
            z = layers.Dense(units=hp.Choice('dense1_units', [32, 64, 128]), kernel_initializer='random_normal')(z)
            z = layers.Dropout(hp.Float('dropout', min_value=0.0, max_value=0.5, step=0.1))(z)
            outputs = layers.Dense(forecast_period)(z)
            model = keras.Model(inputs, outputs)
            model.compile(optimizer=keras.optimizers.Nadam(0.001), loss='mse')
            return model


        tuner = keras_tuner.RandomSearch(build_model, objective='val_loss', max_trials=15,
                                         directory=save_search_path, project_name=search_project_name)
        tuner.search(train_ds, epochs=40, validation_data=validate_ds)
        best_model = tuner.get_best_models(num_models=1)[0]
        best_hyperparameters = tuner.get_best_hyperparameters(num_trials=1)[0]
        best_model.compile(optimizer=keras.optimizers.Nadam(0.001), loss='mse')
        history = best_model.fit(train_ds, epochs=40, verbose=1, validation_data=validate_ds)
        val_loss_avg_mse = best_model.evaluate(validate_ds, verbose=1)
        print(f'Average MSE: {val_loss_avg_mse}')
        print('Best conv1_filters:', best_hyperparameters.get('conv1_filters'))
        print('Best lstm1_units:', best_hyperparameters.get('lstm1_units'))
        print('Best dense1_units:', best_hyperparameters.get('dense1_units'))

    elif forecast_period == 3:
        def build_model(hp):
            inputs = keras.Input(shape=(steps, 17))
            z = layers.Conv1D(filters=hp.Choice('conv1_filters', [16, 32, 64, 128]), kernel_size=3, activation='relu')(inputs)
            z = layers.Conv1D(filters=hp.Choice('conv2_filters', [16, 32, 64, 128]), kernel_size=3, activation='relu')(z)
            z = layers.LSTM(units=hp.Choice('lstm1_units', [16, 32, 64, 128]))(z)
            z = layers.LeakyReLU()(z)
            z = layers.Flatten()(z)
            z = layers.Dense(units=hp.Choice('dense1_units', [32, 64, 128]), kernel_initializer='random_normal')(z)
            outputs = layers.Dense(forecast_period)(z)
            model = keras.Model(inputs, outputs)
            model.compile(optimizer=keras.optimizers.Adam(0.001), loss='mse')
            return model


        tuner = keras_tuner.RandomSearch(build_model, objective='val_loss', max_trials=15,
                                         directory=save_search_path, project_name=search_project_name)
        tuner.search(train_ds, epochs=40, validation_data=validate_ds)
        best_model = tuner.get_best_models(num_models=1)[0]
        best_hyperparameters = tuner.get_best_hyperparameters(num_trials=1)[0]
        best_model.compile(optimizer=keras.optimizers.Adam(0.001), loss='mse')
        history = best_model.fit(train_ds, epochs=40, verbose=1, validation_data=validate_ds)
        val_loss_avg_mse = best_model.evaluate(validate_ds, verbose=1)
        print(f'Average MSE: {val_loss_avg_mse}')
        print('Best conv1_filters:', best_hyperparameters.get('conv1_filters'))
        print('Best conv2_filters:', best_hyperparameters.get('conv2_filters'))
        print('Best lstm1_units:', best_hyperparameters.get('lstm1_units'))
        print('Best dense1_units:', best_hyperparameters.get('dense1_units'))

    elif forecast_period == 5:
        def build_model(hp):
            inputs = keras.Input(shape=(steps, 17))
            z = layers.Conv1D(filters=hp.Choice('conv1_filters', [8, 16, 32, 64, 128]), kernel_size=3, activation='relu')(inputs)
            z = layers.Conv1D(filters=hp.Choice('conv2_filters', [8, 16, 32, 64, 128]), kernel_size=3, activation='relu')(z)
            z = layers.LSTM(units=hp.Choice('lstm1_units', [8, 16, 32, 64, 128]))(z)
            z = layers.LeakyReLU()(z)
            z = layers.Flatten()(z)
            z = layers.Dense(units=hp.Choice('dense1_units', [32, 64, 128]), kernel_initializer='random_normal')(z)
            z = layers.Dense(units=hp.Choice('dense2_units', [32, 64, 128]), kernel_initializer='random_normal')(z)
            z = layers.Dropout(hp.Float('dropout', min_value=0.0, max_value=0.5, step=0.1))(z)
            outputs = layers.Dense(forecast_period)(z)
            model = keras.Model(inputs, outputs)
            model.compile(optimizer=keras.optimizers.Nadam(0.001), loss='mse')
            return model


        tuner = keras_tuner.RandomSearch(build_model, objective='val_loss', max_trials=15,
                                         directory=save_search_path, project_name=search_project_name)
        tuner.search(train_ds, epochs=40, validation_data=validate_ds)
        best_model = tuner.get_best_models(num_models=1)[0]
        best_hyperparameters = tuner.get_best_hyperparameters(num_trials=1)[0]
        best_model.compile(optimizer=keras.optimizers.Nadam(0.001), loss='mse')
        history = best_model.fit(train_ds, epochs=40, verbose=1, validation_data=validate_ds)
        val_loss_avg_mse = best_model.evaluate(validate_ds, verbose=1)
        print(f'Average MSE: {val_loss_avg_mse}')
        print('Best conv1_filters:', best_hyperparameters.get('conv1_filters'))
        print('Best conv2_filters:', best_hyperparameters.get('conv2_filters'))
        print('Best lstm1_units:', best_hyperparameters.get('lstm1_units'))
        print('Best dense1_units:', best_hyperparameters.get('dense1_units'))
        print('Best dense2_units:', best_hyperparameters.get('dense2_units'))

    elif forecast_period == 7:
        def build_model(hp):
            inputs = keras.Input(shape=(steps, 17))
            z = layers.Conv1D(filters=hp.Choice('conv1_filters', [64, 128]), kernel_size=3, activation='relu')(inputs)
            # z = layers.BatchNormalization()(z)
            # z = layers.MaxPooling1D(pool_size=2)(z)
            z = layers.Conv1D(filters=hp.Choice('conv2_filters', [16, 64]), kernel_size=3, activation='relu')(z)
            z = layers.Conv1D(filters=hp.Choice('conv3_filters', [16, 32]), kernel_size=3, activation='relu')(z)
            z = layers.LSTM(units=hp.Choice('lstm1_units', [128]))(z)
            z = layers.LeakyReLU()(z)
            z = layers.Flatten()(z)
            z = layers.Dense(units=hp.Choice('dense1_units', [32]), kernel_initializer='random_normal')(z)
            z = layers.Dense(units=hp.Choice('dense2_units', [64]), kernel_initializer='random_normal')(z)
            outputs = layers.Dense(forecast_period)(z)
            model = keras.Model(inputs, outputs)
            model.compile(optimizer=keras.optimizers.Nadam(0.001), loss='mse')
            return model


        tuner = keras_tuner.RandomSearch(build_model, objective='val_loss', max_trials=15,
                                         directory=save_search_path, project_name=search_project_name)
        tuner.search(train_ds, epochs=40, validation_data=validate_ds)
        best_model = tuner.get_best_models(num_models=1)[0]
        best_hyperparameters = tuner.get_best_hyperparameters(num_trials=1)[0]
        best_model.compile(optimizer=keras.optimizers.Nadam(0.001), loss='mse')
        history = best_model.fit(train_ds, epochs=40, verbose=1, validation_data=validate_ds)
        val_loss_avg_mse = best_model.evaluate(validate_ds, verbose=1)
        print(f'Average MSE: {val_loss_avg_mse}')
        print('Best conv1_filters:', best_hyperparameters.get('conv1_filters'))
        print('Best conv2_filters:', best_hyperparameters.get('conv2_filters'))
        print('Best conv3_filters:', best_hyperparameters.get('conv3_filters'))
        print('Best lstm1_units:', best_hyperparameters.get('lstm1_units'))
        print('Best dense1_units:', best_hyperparameters.get('dense1_units'))
        print('Best dense2_units:', best_hyperparameters.get('dense2_units'))



best_model.save(save_results_path / 'best_model.h5')
with open(save_results_path / 'best_hyperparameters.pkl', 'wb') as f:
    pickle.dump(best_hyperparameters, f)
with open(save_results_path / 'history.pkl', 'wb') as f:
    pickle.dump(history.history, f)

title_size_xy = 14 
number_size_xy = 11 
label_size = 11
text_size = 11  

history_dict = history.history 
train_loss = history_dict['loss'] 
validate_loss = history_dict['val_loss'] 
plt.figure()
plt.plot(range(40), train_loss, label='train_loss')
plt.plot(range(40), validate_loss, label='val_loss')  
plt.legend() 
plt.xlabel('epochs')
plt.ylabel('loss')
plt.savefig(save_results_path / f'loss.png', bbox_inches='tight')
plt.close()



y_train_predict_0 = best_model.predict(x_train)
y_train_predict_1 = y_train_predict_0 * w_std + w_mean
y_train_predict = np.expand_dims(y_train_predict_1, axis=-1)
y_train = y_train * w_std + w_mean
xy_max = int(np.max([np.max(y_train_predict_1), np.max(y_train)])) + 1000
nse_scores, rmse_scores = [], []
for p in range(forecast_period):
    rmse_score = mean_squared_error(y_train_predict[:, p], y_train[:, p], squared=False)
    rmse_scores.append(rmse_score)
    y_mean = np.mean(y_train[:, p]) 
    y_mean_arr = np.full((y_train[:, p].shape[0], 1), y_mean)
    ss_res = mean_squared_error(y_train[:, p], y_train_predict[:, p])  
    ss_tot = mean_squared_error(y_train[:, p], y_mean_arr) 
    nse_score = 1 - ss_res / ss_tot
    nse_scores.append(nse_score)
rmse_scores = [round(score, 1) for score in rmse_scores]
nse_scores = [round(score, 3) for score in nse_scores]
for p, rmse_score in enumerate(rmse_scores):
    print(f"rmse {p + 1}: {rmse_score}") 
for p, nse_score in enumerate(nse_scores):
    print(f"nse {p + 1}: {nse_score}")
for o in range(forecast_period):
    y_train_predict_o = y_train_predict[:, o, :]
    y_train_predict_o_1d = np.squeeze(y_train_predict_o)
    y_train_o = y_train[:, o, :]
    y_train_o_1d = np.squeeze(y_train_o)
    plt.figure(figsize=(5, 5), dpi=300)
    plt.grid(True, color='lightgray', linewidth=0.3, zorder=0)
    parameter = np.polyfit(y_train_o_1d, y_train_predict_o_1d, 1)
    x0 = x00 = [0, np.max(y_train_o_1d)] 
    y0 = [parameter[0] * x0[0] + parameter[1], parameter[0] * x0[1] + parameter[1]] 
    y00 = x00 
    plt.scatter(y_train_o_1d, y_train_predict_o_1d, color='cornflowerblue', s=2, label='Flow', zorder=2)
    plt.plot(x0, y0, linewidth=1, color='g', linestyle='--', label='Trend line', zorder=1)
    plt.plot(x00, y00, linewidth=0.8, color='gray', linestyle='-', label='y=x', zorder=1) 
    plt.legend(fontsize=label_size, loc='upper left') 
    k = '%.3f' % parameter[0] 
    b = '%.3f' % parameter[1] 
    if float(b) <= 0:
        strname = "y=" + k + 'x' + b
    else:
        strname = "y=" + k + 'x+' + b
    r = np.corrcoef(y_train_o_1d, y_train_predict_o_1d)[0, 1]
    r2 = r ** 2
    r2_text = '%.3f' % r2 
    r2_text = 'R =' + r2_text
    rmse_train_predict = 'rmse = ' + str(rmse_scores[o])
    nse_train_predict = 'nse = ' + str(nse_scores[o])
    plt.text(0.6 * xy_max, 0.20 * xy_max, strname, fontsize=text_size)
    plt.text(0.6 * xy_max, 0.145 * xy_max, r2_text, fontsize=text_size)
    plt.text(0.63 * xy_max, 0.17 * xy_max, 2, fontsize=text_size * 0.55)
    plt.text(0.6 * xy_max, 0.10 * xy_max, rmse_train_predict, fontsize=text_size)
    plt.text(0.6 * xy_max, 0.05 * xy_max, nse_train_predict, fontsize=text_size)
    plt.xlim(0, xy_max)
    plt.ylim(0, xy_max)
    plt.tick_params(labelsize=number_size_xy)
    plt.tick_params(pad=5) 
    plt.xlabel('True flow (m$^3$/s)', fontname='Times New Roman', fontsize=title_size_xy)
    plt.ylabel('Predict flow (m$^3$/s)', fontname='Times New Roman', fontsize=title_size_xy)
    plt.rc('font', family='Times New Roman') 
    ax = plt.gca()
    ax.xaxis.set_major_locator(ticker.MultipleLocator(2500))
    plt.gcf().subplots_adjust(left=0.2, right=0.8, top=0.8, bottom=0.2) 
    plt.savefig(save_results_path / f'训练集评价{o}.png', bbox_inches='tight')
    plt.close()



y_predict_0 = best_model.predict(x_test)
y_predict_1 = y_predict_0 * w_std + w_mean
y_predict = np.expand_dims(y_predict_1, axis=-1)
y_test = y_test * w_std + w_mean
xy_max = int(np.max([np.max(y_predict_1), np.max(y_test)]))+1000
nse_scores, rmse_scores = [], []
for p in range(forecast_period):
    rmse_score = mean_squared_error(y_predict[:, p], y_test[:, p], squared=False)
    rmse_scores.append(rmse_score)
    y_mean = np.mean(y_test[:, p])
    y_mean_arr = np.full((y_test[:, p].shape[0], 1), y_mean)
    ss_res = mean_squared_error(y_test[:, p], y_predict[:, p])
    ss_tot = mean_squared_error(y_test[:, p], y_mean_arr)
    nse_score = 1 - ss_res / ss_tot
    nse_scores.append(nse_score)
rmse_scores = [round(score, 1) for score in rmse_scores]
nse_scores = [round(score, 3) for score in nse_scores]
for p, rmse_score in enumerate(rmse_scores):
    print(f"rmse {p + 1}: {rmse_score}")
for p, nse_score in enumerate(nse_scores):
    print(f"nse {p + 1}: {nse_score}")
for m in range(forecast_period):
    y_predict_m = y_predict[:, m, :]
    y_predict_m_1d = np.squeeze(y_predict_m)
    y_test_m = y_test[:, m, :]
    y_test_m_1d = np.squeeze(y_test_m)
    plt.figure(figsize=(5, 5), dpi=300)
    plt.grid(True, color='lightgray', linewidth=0.3, zorder=0)
    parameter = np.polyfit(y_test_m_1d, y_predict_m_1d, 1)
    x0 = x00 = [0, np.max(y_test_m_1d)]
    y0 = [parameter[0] * x0[0] + parameter[1], parameter[0] * x0[1] + parameter[1]]
    y00 = x00
    plt.scatter(y_test_m_1d, y_predict_m_1d, color='cornflowerblue', s=2, label='Flow', zorder=2)
    plt.plot(x0, y0, linewidth=1, color='g', linestyle='--', label='Trend line', zorder=1)
    plt.plot(x00, y00, linewidth=0.8, color='gray', linestyle='-', label='y=x', zorder=1)
    plt.legend(fontsize=label_size, loc='upper left')
    k = '%.3f' % parameter[0]
    b = '%.3f' % parameter[1]
    if float(b) <= 0:
        strname = "y=" + k + 'x' + b
    else:
        strname = "y=" + k + 'x+' + b
    r = np.corrcoef(y_test_m_1d, y_predict_m_1d)[0, 1]
    r2 = r**2
    r2_text = '%.3f' % r2
    r2_text = 'R ='+r2_text
    rmse_predict = 'rmse = ' + str(rmse_scores[m])
    nse_predict = 'nse = ' + str(nse_scores[m])
    plt.text(0.6 * xy_max, 0.20 * xy_max, strname, fontsize=text_size)
    plt.text(0.6 * xy_max, 0.145 * xy_max, r2_text, fontsize=text_size)
    plt.text(0.63 * xy_max, 0.17 * xy_max, 2, fontsize=text_size * 0.55)
    plt.text(0.6 * xy_max, 0.10 * xy_max, rmse_predict, fontsize=text_size)
    plt.text(0.6 * xy_max, 0.05 * xy_max, nse_predict, fontsize=text_size)
    plt.xlim(0, xy_max)
    plt.ylim(0, xy_max)
    plt.tick_params(labelsize=number_size_xy)
    plt.tick_params(pad=5)
    plt.xlabel('True flow (m$^3$/s)', fontname='Times New Roman', fontsize=title_size_xy)
    plt.ylabel('Predict flow (m$^3$/s)', fontname='Times New Roman', fontsize=title_size_xy)
    plt.rc('font', family='Times New Roman')
    ax = plt.gca()
    ax.xaxis.set_major_locator(ticker.MultipleLocator(2500))
    plt.gcf().subplots_adjust(left=0.2, right=0.8, top=0.8, bottom=0.2)
    plt.savefig(save_results_path/f'测试集评价{m}.png', bbox_inches='tight')
    plt.close()



for n in range(forecast_period):
    y_predict_n = y_predict[:, n, :]
    y_predict_n_1d = np.squeeze(y_predict_n)
    y_test_n = y_test[:, n, :]
    y_test_n_1d = np.squeeze(y_test_n)

    fig = plt.figure(figsize=(10, 5), dpi=300)
    axes1 = fig.add_subplot(111)
    axes2 = axes1.twinx()
    axes2.yaxis.tick_right()
    axes2.yaxis.set_label_position('right')
    scatter_actual = axes1.plot(np.arange(len(y_test_n_1d)) + n, y_test_n_1d, 'o', color='cornflowerblue', label='Actuality', markersize=1)
    scatter_predict = axes1.plot(np.arange(len(y_predict_n_1d)) + n, y_predict_n_1d, 'ro', label='Prediction', markersize=1)
    bar_rainfall = axes2.bar(np.arange(len(sum_rainfall_test)), sum_rainfall_test,  width=0.5, color='lightgray', label='Rainfall')
    handles = [scatter_actual[0], scatter_predict[0], bar_rainfall]
    labels = ['Actual', 'Predict', 'Rainfall']

    axes1.set_xlabel('Number of rainfall/flow records from 2005 to 2007_'+str(n+1), fontname='Times New Roman', fontsize=title_size_xy)
    axes1.set_ylabel('Flow (m$^3$/s)', fontname='Times New Roman', fontsize=title_size_xy)
    axes1.tick_params(labelsize=number_size_xy)
    if forecast_period == 1:
        axes1.set_xlim(-20, 620)
        axes1.set_xticks(np.arange(0, 601, 50))
    elif forecast_period == 3:
        axes1.set_xlim(-20, 610)
        axes1.set_xticks(np.arange(0, 601, 50))
    elif forecast_period == 5:
        axes1.set_xlim(-15, 580)
        axes1.set_xticks(np.arange(0, 580, 50))
    elif forecast_period == 7:
        axes1.set_xlim(-20, 560)
        axes1.set_xticks(np.arange(0, 551, 50))
    axes1.set_ylim(0, 30000)
    axes1.set_yticks(np.arange(0, 30001, 5000))

    axes2.set_ylabel('Rainfall (mm/48h)', fontname='Times New Roman', fontsize=title_size_xy)
    axes2.tick_params(labelsize=number_size_xy)
    axes2.set_ylim(400, 0)
    axes2.set_yticks(np.arange(400, -1, -50))

    plt.legend(handles, labels, fontsize=label_size, loc='center left')
    plt.gcf().subplots_adjust(left=0.2, right=0.8, top=0.8, bottom=0.2)
    plt.savefig(save_results_path/f'测试散点图{n}.png')
    plt.close()



explainer = shap.GradientExplainer(best_model, xx_standardization)
shap_values_0 = explainer.shap_values(xx_standardization)
shap_values_abs = np.abs(np.array(shap_values_0, dtype=np.float32))
sum_shap_values_axis1 = np.sum(shap_values_abs, axis=1)
sum_shap_values_axis2 = np.sum(shap_values_abs, axis=2)
workbook = Workbook()
for i in range(forecast_period):
    sheet = workbook.create_sheet(title=f'Sheet{i + 1}')
    matrix1 = sum_shap_values_axis1[i]
    df1 = pd.DataFrame(matrix1)
    for r in dataframe_to_rows(df1, index=False, header=False):
        sheet.append(r)
workbook.save(save_results_path / 'sum_shap_values_axis1.xlsx')
workbook.close()

workbook = Workbook()
for i in range(forecast_period):
    sheet = workbook.create_sheet(title=f'Sheet{i + 1}')
    matrix2 = sum_shap_values_axis2[i]
    df2 = pd.DataFrame(matrix2)
    for r in dataframe_to_rows(df2, index=False, header=False):
        sheet.append(r)
workbook.save(save_results_path / 'sum_shap_values_axis2.xlsx')
workbook.close()
